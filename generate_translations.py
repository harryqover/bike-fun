import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Translations"

# Styling
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="0057FF", end_color="0057FF", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["Key", "en", "fr", "nl", "de"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Translation data
translations = [
    # Loading
    ("bf.loading.title", "Analyzing your profile...", "Analyse de votre profil...", "Uw profiel analyseren...", "Ihr Profil wird analysiert..."),
    
    # Bike Type
    ("bf.bike.type.label", "Bike type", "Type de vélo", "Type fiets", "Fahrradtyp"),
    ("bf.bike.type.regular", "Regular bike", "Vélo classique", "Gewone fiets", "Normales Fahrrad"),
    ("bf.bike.type.regular_ebike", "Regular e-bike", "Vélo électrique classique", "Gewone e-bike", "Normales E-Bike"),
    ("bf.bike.type.mountain", "Mountain bike", "VTT", "Mountainbike", "Mountainbike"),
    ("bf.bike.type.mountain_ebike", "Mountain e-bike", "VTT électrique", "Elektrische mountainbike", "E-Mountainbike"),
    ("bf.bike.type.race", "Race bike", "Vélo de course", "Racefiets", "Rennrad"),
    ("bf.bike.type.race_ebike", "Race e-bike", "Vélo de course électrique", "Elektrische racefiets", "E-Rennrad"),
    ("bf.bike.type.cargo", "Cargo bike", "Vélo cargo", "Bakfiets", "Lastenrad"),
    ("bf.bike.type.cargo_ebike", "Cargo e-bike", "Vélo cargo électrique", "Elektrische bakfiets", "E-Lastenrad"),
    
    # Bike Value & Anti-theft
    ("bf.bike.value.label", "Bike value (accessories included)", "Valeur du vélo (accessoires inclus)", "Waarde van de fiets (accessoires inbegrepen)", "Fahrradwert (inkl. Zubehör)"),
    ("bf.bike.antiTheft.label", "Anti-theft measure", "Mesure antivol", "Antidiefstal", "Diebstahlschutz"),
    ("bf.bike.antiTheft.none", "None", "Aucune", "Geen", "Keine"),
    ("bf.bike.antiTheft.gps", "GPS Tracker", "Traceur GPS", "GPS-tracker", "GPS-Tracker"),
    
    # Postal Code & Button
    ("bf.policyholder.address.zip.label", "Postal code", "Code postal", "Postcode", "Postleitzahl"),
    ("bf.button.seePrice", "See prices", "Voir les prix", "Bekijk prijzen", "Preise anzeigen"),
    
    # Coverage
    ("bf.coverage.title", "Choose your coverage", "Sélection de la couverture", "Kies uw dekking", "Wählen Sie Ihren Schutz"),
    ("bf.coverage.package.theft", "THEFT + ASSISTANCE", "VOL UNIQUEMENT", "ALLEEN DIEFSTAL", "NUR DIEBSTAHL"),
    ("bf.coverage.package.damage", "DAMAGE + ASSISTANCE", "DOMMAGES UNIQUEMENT", "ALLEEN SCHADE", "NUR SCHADEN"),
    ("bf.coverage.package.comprehensive", "OMNIUM", "VOL + DOMMAGES + ASSISTANCE", "DIEFSTAL + SCHADE + ASSISTENTIE", "DIEBSTAHL + SCHADEN + ASSISTENZ"),
    ("bf.coverage.theft", "Theft of your bike", "Vol de votre vélo", "Diefstal van uw fiets", "Diebstahl Ihres Fahrrads"),
    ("bf.coverage.assistance", "Assistance 24/7", "Assistance 24/7", "Assistentie 24/7", "Assistance 24/7"),
    ("bf.coverage.damage", "Material damage", "Dommages matériels", "Materiële schade", "Sachschäden"),
    
    # Bike Details Section
    ("bf.bike.title", "About your bike", "À propos de votre vélo", "Over uw fiets", "Über Ihr Fahrrad"),
    ("bf.bike.make.label", "Bike make", "Marque du vélo", "Merk van de fiets", "Fahrradmarke"),
    ("bf.bike.make.placeholder", "e.g. Canyon", "p.ex. Canyon", "bijv. Canyon", "z.B. Canyon"),
    ("bf.bike.model.label", "Bike model", "Modèle du vélo", "Model van de fiets", "Fahrradmodell"),
    ("bf.bike.model.placeholder", "e.g. Spectral", "p.ex. Spectral", "bijv. Spectral", "z.B. Spectral"),
    ("bf.bike.serialNumber.label", "Frame/Serial number", "Numéro de série/cadre", "Frame-/serienummer", "Rahmen-/Seriennummer"),
    ("bf.bike.serialNumber.placeholder", "Enter the serial number", "Entrez le numéro de série", "Voer het serienummer in", "Seriennummer eingeben"),
    ("bf.bike.purchaseDate.label", "Purchase date", "Date d'achat", "Aankoopdatum", "Kaufdatum"),
    ("bf.bike.newBike.label", "Is this a new bike?", "S'agit-il d'un vélo neuf ?", "Is dit een nieuwe fiets?", "Ist das Fahrrad neu?"),
    ("bf.bike.newBike.yes", "Yes - New", "Oui - Neuf", "Ja - Nieuw", "Ja - Neu"),
    ("bf.bike.newBike.no", "No - Second-hand", "Non - Occasion", "Nee - Tweedehands", "Nein - Gebraucht"),
    
    # Policyholder
    ("bf.policyholder.title", "Policyholder information", "Informations du souscripteur", "Informatie verzekeringsnemer", "Versicherungsnehmer"),
    ("bf.policyholder.entity.personal", "Private person", "Particulier", "Particulier", "Privatperson"),
    ("bf.policyholder.entity.company", "Company", "Entreprise", "Bedrijf", "Unternehmen"),
    ("bf.policyholder.company.name.label", "Company name", "Nom de l'entreprise", "Bedrijfsnaam", "Firmenname"),
    ("bf.policyholder.company.name.placeholder", "Enter company name", "Entrez le nom de l'entreprise", "Voer bedrijfsnaam in", "Firmennamen eingeben"),
    ("bf.policyholder.company.vat.label", "VAT number", "Numéro de TVA", "BTW-nummer", "USt-IdNr."),
    ("bf.policyholder.company.vat.placeholder", "BE0123456789", "BE0123456789", "BE0123456789", "DE123456789"),
    ("bf.policyholder.firstName.label", "First name", "Prénom", "Voornaam", "Vorname"),
    ("bf.policyholder.firstName.placeholder", "John", "Jean", "Jan", "Max"),
    ("bf.policyholder.lastName.label", "Last name", "Nom", "Achternaam", "Nachname"),
    ("bf.policyholder.lastName.placeholder", "Doe", "Dupont", "Jansen", "Mustermann"),
    ("bf.policyholder.email.label", "Email", "Adresse e-mail", "E-mailadres", "E-Mail"),
    ("bf.policyholder.email.placeholder", "john.doe@example.com", "jean.dupont@exemple.fr", "jan.jansen@voorbeeld.nl", "max.mustermann@beispiel.de"),
    ("bf.policyholder.phone.label", "Phone number", "Numéro de téléphone", "Telefoonnummer", "Telefonnummer"),
    ("bf.policyholder.phone.placeholder", "+33 1 23 45 67 89", "+33 1 23 45 67 89", "+31 6 12 34 56 78", "+49 30 12345678"),
    ("bf.policyholder.birthdate.label", "Date of birth", "Date de naissance", "Geboortedatum", "Geburtsdatum"),
    ("bf.policyholder.address.street.label", "Street and number", "Rue et numéro", "Straat en huisnummer", "Straße und Hausnummer"),
    ("bf.policyholder.address.street.placeholder", "Main Street 123", "Rue Principale 123", "Hoofdstraat 123", "Hauptstraße 123"),
    ("bf.policyholder.address.city.label", "City", "Ville", "Stad", "Stadt"),
    ("bf.policyholder.address.city.placeholder", "Brussels", "Bruxelles", "Brussel", "Berlin"),
    
    # Start Date & Legal
    ("bf.startDate.title", "Insurance start date", "Startdatum verzekering", "Datum de début de l'assurance", "Versicherungsbeginn"),
    ("bf.startDate.setNow", "I want to set a start date now", "Je souhaite définir une date de début maintenant", "Ik wil nu een startdatum instellen", "Ich möchte jetzt ein Startdatum festlegen"),
    ("bf.terms.acceptance.label", "I confirm that I have read and accepted the electronic communication agreement", "Je confirme avoir lu et accepté l'accord de communication électronique", "Ik bevestig dat ik de elektronische communicatieovereenkomst heb gelezen en geaccepteerd", "Ich bestätige, dass ich die elektronische Kommunikationsvereinbarung gelesen und akzeptiert habe"),
    ("bf.terms.important.label", "I agree to receive information from Qover by email including PDF attachments", "J'accepte de recevoir des informations de Qover par e-mail, y compris des pièces jointes PDF", "Ik ga akkoord met het ontvangen van informatie van Qover per e-mail inclusief PDF-bijlagen", "Ich stimme zu, Informationen von Qover per E-Mail einschließlich PDF-Anhängen zu erhalten"),
    ("bf.terms.eligibility.label", "I confirm that I have read and accepted the General Terms and Conditions and IPID", "Je confirme avoir lu et accepté les Conditions Générales et l'IPID", "Ik bevestig dat ik de Algemene Voorwaarden en het IPID heb gelezen en geaccepteerd", "Ich bestätige, dass ich die AGB und das IPID gelesen und akzeptiert habe"),
    ("bf.terms.privacy.label", "I have read the Privacy Policy and understand how my personal data will be processed", "J'ai lu la Politique de Confidentialité et je comprends comment mes données personnelles seront traitées", "Ik heb het Privacybeleid gelezen en begrijp hoe mijn persoonlijke gegevens worden verwerkt", "Ich habe die Datenschutzrichtlinie gelesen und verstehe, wie meine persönlichen Daten verarbeitet werden"),
    ("bf.button.submit", "Continue to Payment", "Continuer vers le paiement", "Verder naar betaling", "Weiter zur Zahlung"),
    
    # Sidebar
    ("bf.sidebar.coverage.title", "Your coverage", "Votre couverture", "Uw dekking", "Ihr Schutz"),
    ("bf.sidebar.coverage.vat", "includes VAT", "TVA incluse", "BTW inbegrepen", "inkl. MwSt."),
    ("bf.sidebar.help.title", "Need help with your contract?", "Besoin d'aide avec votre contrat ?", "Hulp nodig bij uw contract?", "Hilfe bei Ihrem Vertrag benötigt?"),
    ("bf.sidebar.help.hours", "Monday to Friday, 9 am to 4 pm.", "Du lundi au vendredi, de 9h à 16h.", "Maandag tot vrijdag, 9.00 tot 16.00 uur.", "Montag bis Freitag, 9 bis 16 Uhr."),
    ("bf.sidebar.help.chat", "Chat", "Chat", "Chat", "Chat"),
    ("bf.sidebar.legal.title", "Legal information", "Informations légales", "Juridische informatie", "Rechtliche Informationen"),
    ("bf.sidebar.legal.termsLink", "Terms and conditions", "Conditions générales", "Algemene voorwaarden", "Allgemeine Geschäftsbedingungen"),
]

# Write translations
for row, (key, en, fr, nl, de) in enumerate(translations, 2):
    ws.cell(row=row, column=1, value=key).border = thin_border
    ws.cell(row=row, column=2, value=en).border = thin_border
    ws.cell(row=row, column=3, value=fr).border = thin_border
    ws.cell(row=row, column=4, value=nl).border = thin_border
    ws.cell(row=row, column=5, value=de).border = thin_border

# Column widths
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 60
ws.column_dimensions['E'].width = 60

# Save
wb.save('/Users/HarryEvrard/Documents/GitHub/bike-fun/bike_flow_translations.xlsx')
print("Created: bike_flow_translations.xlsx")
